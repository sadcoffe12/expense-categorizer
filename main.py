import pandas as pd
import openpyxl
import os
import unicodedata
import re
from collections import Counter
from difflib import SequenceMatcher
from datetime import datetime, timedelta

RULES_FILE = "categorization_rules.csv"
TEMPLATES_FILE = "templates.txt"
# Configuración
HISTORY_FILE = "uncategorized_history.csv"
SIMILARITY_THRESHOLD = 0.8  # 80% de similitud
MAX_HISTORY_AGE_DAYS = 213  # Eliminar entradas más antiguas que 7 meses

# Configuración de Triggers por Período
# Número de repeticiones requeridas para sugerir una regla en cada período
TRIGGER_THIS_SESSION = 5    # Esta sesión: 5 repeticiones
TRIGGER_LAST_MONTH = 3      # Último mes: 3 repeticiones
TRIGGER_LAST_3_MONTHS = 3   # Últimos 3 meses: 3 repeticiones
TRIGGER_LAST_6_MONTHS = 3   # Últimos 6 meses: 3 repeticiones

def normalize_text(text, is_transaction=False):
    """
    Normaliza texto (quitar tildes, símbolos, minúsculas).
    Si is_transaction=True, también elimina ruido bancario (tarjetas, IDs).
    """
    if not isinstance(text, str):
        return str(text) if text is not None else ""

    # 1. Limpieza básica (Minúsculas y acentos)
    text = text.lower().strip()
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                  if unicodedata.category(c) != 'Mn')

    # 2. Limpieza de Ruido Bancario (Solo si es una descripción de gasto)
    if is_transaction:
        # Quita "tarj nro. 1234" o "tarjeta 1234"
        text = re.sub(r'tarj\s?nro\.?\s?\d+', '', text)
        # Quita números largos de 5 o más dígitos (IDs de transacción)
        text = re.sub(r'\d{5,}', '', text)

    # 3. Quitar símbolos y puntuación (mantiene letras, números y espacios)
    # text = re.sub(r'[^\w\s]', ' ', text)
    
    # 4. Colapsar espacios múltiples
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text

def select_column(df, prompt, default_name, template_suggestion=None):
    # Si el template ya trae una columna válida, la usamos sin preguntar
    if template_suggestion and template_suggestion in df.columns:
        print(f"[Template] Usando columna '{template_suggestion}' para {prompt}.")
        return template_suggestion

    # Si no hay sugerencia o no es válida, procedemos con la selección manual
    print(f"\n--- Selección de Columna: {prompt} ---")
    for i, col_name in enumerate(df.columns):
        print(f"  {i+1}. {col_name}")
    
    user_input = input(f"Selecciona número o nombre (Enter para '{default_name}'): ").strip()

    if not user_input:
        return default_name
    
    if user_input.isdigit():
        idx = int(user_input) - 1
        return df, None.columns[idx] if 0 <= idx < len(df.columns) else default_name
    
    return user_input

def load_rules():
    if not os.path.exists(RULES_FILE):
        df_empty = pd.DataFrame(columns=['keyword', 'type', 'category', 'new_description'])
        df_empty.to_csv(RULES_FILE, index=False, encoding="utf-8")
        return []
    
    try:
        # Cargamos y convertimos TODO a string de entrada para evitar floats/NaN
        df_rules = pd.read_csv(RULES_FILE, encoding="utf-8").fillna("")
        
        rules = []
        for _, row in df_rules.iterrows():
            keyword = str(row['keyword']).lower().strip()
            # Si la keyword está vacía (fila accidental en CSV), la saltamos
            if not keyword:
                continue
                
            new_desc = str(row['new_description']).strip()
            # Si está vacío o es un nulo de pandas, pasamos None real
            val_new_desc = new_desc if new_desc not in ["", "nan", "None", "NaN"] else None
            
            rules.append((
                keyword, 
                str(row['type']), 
                str(row['category']), 
                val_new_desc
            ))
        return rules
    except Exception as e:
        print(f"Error al cargar las reglas CSV: {e}")
        return []

# def categorize(df, selected_template=None):
#     print("\n--- Categorizar Registros ---")
    
#     # 1. Identificar columnas (esto DEBE pasar siempre para preparar la estructura)
#     t_source = selected_template.get('SOURCE_COL') if selected_template else None
#     t_type = selected_template.get('TYPE_COL') if selected_template else None
#     t_cat = selected_template.get('CAT_COL') if selected_template else None

#     source_col = select_column(df, "Descripción", "Descripcion", t_source)
#     type_col = select_column(df, "Tipo de Gasto", "Tipo", t_type)
#     category_col = select_column(df, "Categoría", "Categoria", t_cat)

#     # Aseguramos que las columnas de destino existan
#     if type_col not in df.columns:
#         df[type_col] = ""
#     if category_col not in df.columns:
#         df[category_col] = ""
        
#     # Aseguramos tipos string para evitar errores de comparación
#     df[source_col] = df[source_col].astype(str)
#     df[category_col] = df[category_col].astype(str)
#     df[type_col] = df[type_col].astype(str)

#     # 2. Cargar y aplicar reglas (solo si existen)
#     rules = load_rules()
#     if not rules:
#         print("Aviso: No hay reglas definidas en el CSV. Se saltará la categorización automática.")
#     else:
#         categorized_count = 0

#         def assign_category(description):
#             nonlocal categorized_count
#             desc_value = str(description) if pd.notna(description) else ""
#             desc_lower = desc_value.lower()
#             cat_value, type_value = "", ""

#             for keyword, t_val, c_val, new_desc in rules:
#                 if keyword in desc_lower:
#                     categorized_count += 1
#                     cat_value, type_value = c_val, t_val
#                     if new_desc:
#                         desc_value = new_desc
#                         desc_lower = desc_value.lower()
#             return type_value, cat_value, desc_value

#         # Aplicar reglas
#         results = df[source_col].apply(assign_category)
#         results_df = pd.DataFrame(results.tolist(), columns=[type_col, category_col, source_col])
#         df[type_col] = results_df[type_col].values
#         df[category_col] = results_df[category_col].values
#         df[source_col] = results_df[source_col].values
#         print(f"Proceso completado. Se categorizaron {categorized_count} coincidencias.")

#     # 3. Normalización final de nombres de columnas (DEBE pasar siempre)
#     # Esto transforma "descripcion" -> "Descripcion" y "categoria" -> "Categoria"
#     df.columns = [col.replace('_', ' ').capitalize() for col in df.columns]
    
#     return df

def categorize(df, selected_template=None, col_identity=(None, None, None)):
    print("\n--- Categorizar Registros ---")
    
    src, typ, cat = col_identity

    # Si no vienen nombres del template o de la llamada anterior, los pedimos
    if not src or src not in df.columns:
        print(f"Columnas actuales: {list(df.columns)}")
        src = select_column(df, "Descripción", "Descripcion", src)
    if not typ:
        typ = select_column(df, "Tipo de Gasto", "Tipo", typ)
    if not cat:
        cat = select_column(df, "Categoría", "Categoria", cat)

    # Aseguramos existencia de columnas de destino
    if typ not in df.columns: df[typ] = ""
    if cat not in df.columns: df[cat] = ""
        
    df[src] = df[src].astype(str)
    df[cat] = df[cat].astype(str)
    df[typ] = df[typ].astype(str)

    rules = load_rules()
    if not rules:
        print("Aviso: No hay reglas definidas. Saltando categorización automática.")
    else:
        categorized_count = 0
        def assign_category(description):
            nonlocal categorized_count
            desc_value = str(description) if pd.notna(description) else ""
            desc_lower = desc_value.lower()
            cat_v, typ_v = "", ""

            for keyword, t_val, c_val, new_desc in rules:
                if keyword in desc_lower:
                    categorized_count += 1
                    cat_v, typ_v = c_val, t_val
                    if new_desc:
                        desc_value = new_desc
                        desc_lower = desc_value.lower()
            return typ_v, cat_v, desc_value

        results = df[src].apply(assign_category)
        results_df = pd.DataFrame(results.tolist(), columns=[typ, cat, src])
        df[typ], df[cat], df[src] = results_df[typ].values, results_df[cat].values, results_df[src].values
        print(f"Proceso completado. {categorized_count} coincidencias encontradas.")

    # Normalización final de nombres de columnas
    df.columns = [col.replace('_', ' ').capitalize() for col in df.columns]
    
    # Retornamos el DF y los nombres que resultaron de la capitalización
    return df, src.replace('_', ' ').capitalize(), cat.replace('_', ' ').capitalize()

def get_similarity(a, b):
    """Retorna el ratio de similitud entre dos strings."""
    return SequenceMatcher(None, a, b).ratio()

def learn_and_suggest(df, source_col, category_col):
    """Analiza patrones de entradas no categorizadas considerando sesiones pasadas y actuales.
    
    Flujo:
    1. Identifica no categorizadas de ESTA SESIÓN
    2. Busca patrones en: esta sesión, último mes, últimos 3 meses, últimos 6 meses
    3. Si un patrón se repite en cualquier período (según trigger) → sugiere regla
    4. El usuario acepta/rechaza/ingresa manual, y se guarda la regla
    5. Limpia entradas antiguas (> 6 meses)
    6. Actualiza el historial
    
    Triggers configurables:
    - TRIGGER_THIS_SESSION (5): Esta sesión
    - TRIGGER_LAST_MONTH (3): Último mes
    - TRIGGER_LAST_3_MONTHS (3): Últimos 3 meses
    - TRIGGER_LAST_6_MONTHS (3): Últimos 6 meses
    """
    
    # Verificación de seguridad
    if source_col not in df.columns or category_col not in df.columns:
        print(f"Error: No se encontraron las columnas '{source_col}' o '{category_col}' para analizar.")
        print(f"Columnas disponibles: {list(df.columns)}")
        return
    
    # Identifica entradas NO categorizadas de ESTA SESIÓN
    mask_empty = df[category_col].isin(["", "nan", "None", None])
    unassigned = df[mask_empty][source_col].tolist()
    
    if not unassigned:
        print("✅ Todas las entradas han sido categorizadas.")
        return

    # Analiza patrones considerando sesiones pasadas
    patterns_analysis = analyze_patterns_by_period(unassigned)
    
    # Combina todos los patrones encontrados eliminando duplicados
    all_patterns = {
        **patterns_analysis['this_session'],
        **patterns_analysis['last_month'],
        **patterns_analysis['last_3_months'],
        **patterns_analysis['last_6_months']
    }
    
    if not all_patterns:
        print(f"\nℹ️  No hay patrones repetidos en los {len(unassigned)} no categorizados")
        print("   (Se necesitan al menos 3 repeticiones en esta sesión, último mes o últimos 3 meses)")
        # Aun así, guardamos en el historial para análisis futuro
        update_history(unassigned)
        return

    rules = load_rules()
    print("\n--- 🧠 Analizador de Gastos Recurrentes (Análisis Temporal) ---")
    print(f"Se encontraron {len(all_patterns)} patrones repetidos considerando sesiones pasadas.\n")
    
    processed_patterns = set()
    total_patterns = len(all_patterns)
    current_pattern_num = 0
    
    for pattern in all_patterns.keys():
        if pattern in processed_patterns:
            continue
        processed_patterns.add(pattern)
        current_pattern_num += 1
        
        # Mostrar el progreso
        print(f"\n📋 Sugerencia ({current_pattern_num}/{total_patterns})")
        
        # Determina el período donde se detectó
        period_labels = []
        if pattern in patterns_analysis['this_session']:
            count = patterns_analysis['this_session'][pattern]
            period_labels.append(f"Esta sesión ({count}x)")
        if pattern in patterns_analysis['last_month']:
            count = patterns_analysis['last_month'][pattern]
            period_labels.append(f"Último mes ({count}x)")
        if pattern in patterns_analysis['last_3_months']:
            count = patterns_analysis['last_3_months'][pattern]
            period_labels.append(f"Últimos 3 meses ({count}x)")
        if pattern in patterns_analysis['last_6_months']:
            count = patterns_analysis['last_6_months'][pattern]
            period_labels.append(f"Últimos 6 meses ({count}x)")
        
        periods_str = " | ".join(period_labels)
        
        print(f"{'='*70}")
        print(f"Patrón: '{pattern}'")
        print(f"Períodos: {periods_str}")
        
        # Intenta sugerir categoría basada en reglas existentes
        sug_type, sug_cat = guess_category(pattern, rules)
        
        # Flujo de decisión del usuario
        if sug_type:
            print(f"🤖 Sugerencia IA: Tipo={sug_type}, Categoría={sug_cat}")
            confirm = input("¿Aceptar esta categoría? (s/n/omitir): ").lower().strip()
        else:
            confirm = 'n'

        if confirm == 's':
            tipo, cat = sug_type, sug_cat
        elif confirm == 'n':
            print("Ingresa la categorización manual:")
            tipo = input("  Tipo (Fijo/Variable): ").strip()
            cat = input("  Categoría: ").strip()
            if not tipo or not cat:
                print("  ⚠️  Tipo/Categoría vacíos. Saltando este patrón.")
                continue
        else:
            print("  ➜ Omitido.")
            continue

        # Permite renombrar la entrada
        new_desc = input(f"  Renombrar entrada [Enter para '{pattern}']: ").strip()
        final_desc = new_desc if new_desc else pattern
        
        # Guarda la nueva regla
        new_rule = pd.DataFrame([[pattern, tipo, cat, final_desc]], 
                               columns=['keyword', 'type', 'category', 'new_description'])
        new_rule.to_csv(RULES_FILE, mode='a', header=False, index=False, encoding="utf-8")
        print("✅ Regla guardada.")
    
    # Al final, actualiza el historial con las nuevas entradas no categorizadas
    update_history(unassigned)

def update_history(new_list):
    """Mantiene el archivo de memoria histórica de gastos no categorizados con timestamps.
    
    Evita duplicados verificando si la combinación de 'cleaned' y 'date_added' ya existe.
    """
    # Normalizamos con limpieza de ruido bancario
    current_date = datetime.now().strftime("%Y-%m-%d")
    new_data = [
        {
            "original": x, 
            "cleaned": normalize_text(x, is_transaction=True),
            "date_added": current_date
        } for x in new_list
    ]
    df_new = pd.DataFrame(new_data)
    
    if os.path.exists(HISTORY_FILE):
        df_old = pd.read_csv(HISTORY_FILE)
        # Asegurar que la columna 'date_added' existe en df_old (por compatibilidad)
        if 'date_added' not in df_old.columns:
            df_old['date_added'] = datetime.now().strftime("%Y-%m-%d")
        
        # Eliminar duplicados: comparar por 'cleaned' Y 'date_added'
        # Crear una clave única para cada entrada
        df_old['_key'] = df_old['cleaned'] + '|' + df_old['date_added'].astype(str)
        df_new['_key'] = df_new['cleaned'] + '|' + df_new['date_added'].astype(str)
        
        # Identificar qué entradas de df_new ya existen en df_old
        new_keys = set(df_new['_key'])
        old_keys = set(df_old['_key'])
        duplicated_keys = new_keys & old_keys
        
        # Solo mantener las nuevas que NO están duplicadas
        df_new_filtered = df_new[~df_new['_key'].isin(duplicated_keys)].copy()
        
        # Limpiar las claves temporales
        df_old = df_old.drop(columns=['_key'])
        df_new_filtered = df_new_filtered.drop(columns=['_key'])
        
        # Si hay nuevas entradas sin duplicar, agregarlas
        if not df_new_filtered.empty:
            df_final = pd.concat([df_old, df_new_filtered], ignore_index=True)
            duplicated_count = len(duplicated_keys)
            if duplicated_count > 0:
                print(f"ℹ️  {duplicated_count} entrada(s) ya existía(n) en el historial. No se duplicaron.")
        else:
            df_final = df_old
            print(f"ℹ️  Todas las entradas ({len(new_list)}) ya existían en el historial. No se agregaron duplicados.")
    else:
        df_final = df_new
    
    # Limpiar entradas antiguas
    df_final = clean_old_entries(df_final)
    
    df_final.to_csv(HISTORY_FILE, index=False, encoding="utf-8")
    return df_final

def clean_old_entries(df_history):
    """Elimina entradas más antiguas que MAX_HISTORY_AGE_DAYS."""
    if df_history.empty or 'date_added' not in df_history.columns:
        return df_history
    
    df_history['date_added'] = pd.to_datetime(df_history['date_added'], errors='coerce')
    cutoff_date = datetime.now() - timedelta(days=MAX_HISTORY_AGE_DAYS)
    
    df_clean = df_history[df_history['date_added'] >= cutoff_date].copy()
    
    removed_count = len(df_history) - len(df_clean)
    if removed_count > 0:
        print(f"ℹ️  Se eliminaron {removed_count} entradas con más de {MAX_HISTORY_AGE_DAYS} días de antigüedad.")
    
    return df_clean

def guess_category(cleaned_desc, rules):
    """Intenta predecir categoría basándose en reglas existentes."""
    best_score = 0
    best_match = (None, None) # (Tipo, Categoria)
    
    for keyword, t_val, c_val, _ in rules:
        # Comparamos la keyword de la regla con el texto actual
        score = SequenceMatcher(None, cleaned_desc, keyword).ratio()
        if score > best_score:
            best_score = score
            best_match = (t_val, c_val)
            
    # Si la confianza es mayor al 60%, sugerimos
    return best_match if best_score > 0.6 else (None, None)

def analyze_patterns_by_period(current_unassigned):
    """Analiza patrones en sesiones pasadas y actuales considerando períodos de tiempo.
    
    Verifica patrones en 4 períodos con triggers configurables:
    1. Esta sesión (TRIGGER_THIS_SESSION = 5 repeticiones)
    2. Último mes (TRIGGER_LAST_MONTH = 3 repeticiones)
    3. Últimos 3 meses (TRIGGER_LAST_3_MONTHS = 3 repeticiones)
    4. Últimos 6 meses (TRIGGER_LAST_6_MONTHS = 3 repeticiones)
    
    Retorna: dict con patrones y su contexto temporal
    """
    
    if not os.path.exists(HISTORY_FILE):
        # Si no hay historial, solo analizar esta sesión
        new_unassigned = [{"cleaned": normalize_text(x, is_transaction=True)} for x in current_unassigned]
        df_new = pd.DataFrame(new_unassigned)
        counts = df_new['cleaned'].value_counts()
        return {
            'this_session': counts[counts >= TRIGGER_THIS_SESSION].to_dict(),
            'last_month': {},
            'last_3_months': {},
            'last_6_months': {}
        }
    
    try:
        df_history = pd.read_csv(HISTORY_FILE)
        
        # Asegurar que la columna 'date_added' existe
        if 'date_added' not in df_history.columns:
            df_history['date_added'] = datetime.now().strftime("%Y-%m-%d")
        
        df_history['date_added'] = pd.to_datetime(df_history['date_added'], errors='coerce')
        now = datetime.now()
        
        # 1. Patrones en ESTA SESIÓN
        new_unassigned = [{"cleaned": normalize_text(x, is_transaction=True)} for x in current_unassigned]
        df_new = pd.DataFrame(new_unassigned)
        counts_this_session = df_new['cleaned'].value_counts()
        patterns_this_session = counts_this_session[counts_this_session >= TRIGGER_THIS_SESSION].to_dict()
        
        # 2. Patrones en el ÚLTIMO MES
        one_month_ago = now - timedelta(days=30)
        df_last_month = df_history[df_history['date_added'] >= one_month_ago]
        counts_month = df_last_month['cleaned'].value_counts()
        patterns_last_month = counts_month[counts_month >= TRIGGER_LAST_MONTH].to_dict()
        
        # 3. Patrones en los ÚLTIMOS 3 MESES
        three_months_ago = now - timedelta(days=90)
        df_last_3m = df_history[df_history['date_added'] >= three_months_ago]
        counts_3m = df_last_3m['cleaned'].value_counts()
        patterns_last_3m = counts_3m[counts_3m >= TRIGGER_LAST_3_MONTHS].to_dict()
        
        # 4. Patrones en los ÚLTIMOS 6 MESES
        six_months_ago = now - timedelta(days=180)
        df_last_6m = df_history[df_history['date_added'] >= six_months_ago]
        counts_6m = df_last_6m['cleaned'].value_counts()
        patterns_last_6m = counts_6m[counts_6m >= TRIGGER_LAST_6_MONTHS].to_dict()
        
        return {
            'this_session': patterns_this_session,
            'last_month': patterns_last_month,
            'last_3_months': patterns_last_3m,
            'last_6_months': patterns_last_6m
        }
        
    except Exception as e:
        print(f"Error al analizar patrones históricos: {e}")
        return {
            'this_session': {},
            'last_month': {},
            'last_3_months': {},
            'last_6_months': {}
        }

def apply_format(df, file_path):
    #Aplica un template para limpiar un archivo, y luego elimina/añade columnas/filas 
    # en el orden correcto. Devuelve solo el DataFrame en memoria (no guarda archivo).

    print("--- Aplicación de un template de formato de limpieza ---")
    
    if not os.path.exists(TEMPLATES_FILE):
        print("No se encontraron templates de formato. Por favor, crea uno primero.")
        return df, None

    # --- (La sección de lectura y selección de templates no cambia) ---
    templates = {}
    try:
        with open(TEMPLATES_FILE, "r") as f:
            content = f.read().strip()
            if not content:
                print("El archivo de templates está vacío.")
                return df, None, None
            
            lines = content.split("---")
            for line_block in lines:
                line_block = line_block.strip()
                if not line_block: 
                    continue
                
                current_template = {}
                for line in line_block.split('\n'):
                    line = line.strip()
                    if ':' in line:
                        try:
                            key, value = line.split(": ", 1)
                            key = key.strip()
                            value = value.strip()
                            
                            keys_to_normalize = ['COLS_TO_ADD', 'ORDERED_COLS', 'SOURCE_COL', 'TYPE_COL', 'CAT_COL']
                            if key in keys_to_normalize:
                                if ',' in value:
                                    value = ", ".join([normalize_text(v) for v in value.split(',')])
                                else:
                                    value = normalize_text(value)
                            
                            current_template[key] = value
                        except ValueError:
                            continue
                
                if 'TEMPLATE_NAME' in current_template:
                    templates[current_template['TEMPLATE_NAME']] = current_template
                    
    except Exception as e:
        print(f"Error al leer los templates: {e}")
        return df, None

    if not templates:
        print("No se encontraron templates de formato.")
        return df, None

    print("Templates de formato disponibles:")
    for i, name in enumerate(templates.keys()):
        print(f"  {i+1}. {name}")
    
    try:
        choice = int(input("Selecciona el número del template a aplicar: "))
        selected_name = list(templates.keys())[choice - 1]
        selected_template = templates[selected_name]
    except (ValueError, IndexError):
        print("Selección no válida.")
        return df, None

    # --- Comienza la lógica de aplicación ---
    try:
        # 1. Leer los datos del rango especificado
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        header_row = int(selected_template['HEADER_ROW'])
        start_row = int(selected_template['START_ROW'])
        start_col_letter = selected_template['START_COL']
        end_col_letter = selected_template['END_COL']
        
        start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)
        end_col_idx = openpyxl.utils.column_index_from_string(end_col_letter)

        data = []
        for r in range(header_row, ws.max_row + 1):
            row_data = [ws.cell(row=r, column=c).value for c in range(start_col_idx, end_col_idx + 1)]
            data.append(row_data)

        # 2. Crear el DataFrame en memoria
        if not data:
            print("No se encontraron datos en el rango especificado.")
            return df, None
            
        df_new = pd.DataFrame(data[1:], columns=data[0])
        
        # Limpieza de encabezados
        df_new.columns = [normalize_text(col) for col in df_new.columns]
        print("Encabezados normalizados (sin tildes ni símbolos).")

        df_new.dropna(how='all', inplace=True)

        # 3. Eliminar columnas especificadas
        if 'COLS_TO_DROP' in selected_template and selected_template['COLS_TO_DROP']:
            cols_to_drop = [col.strip() for col in selected_template['COLS_TO_DROP'].split(',') if col.strip()]
            headers_map = {openpyxl.utils.get_column_letter(start_col_idx + i): header for i, header in enumerate(data[0])}
            cols_to_drop_names = [headers_map[letter] for letter in cols_to_drop if letter in headers_map]
            cols_to_drop_names_existing = [name for name in cols_to_drop_names if name in df_new.columns]
            
            df_new.drop(columns=cols_to_drop_names_existing, inplace=True, errors='ignore')
            if cols_to_drop_names_existing:
                print(f"Columnas eliminadas: {', '.join(cols_to_drop_names_existing)}")

        # 4. Eliminar filas especificadas
        if 'ROWS_TO_DROP' in selected_template and selected_template['ROWS_TO_DROP']:
            rows_to_drop_str = [r.strip() for r in selected_template['ROWS_TO_DROP'].split(',') if r.strip()]
            rows_to_drop_indices = [int(r_str) - start_row for r_str in rows_to_drop_str]
            rows_to_drop_indices = [idx for idx in rows_to_drop_indices if 0 <= idx < len(df_new)]
            
            df_new.drop(index=rows_to_drop_indices, inplace=True, errors='ignore')
            if rows_to_drop_indices:
                print(f"Se eliminaron {len(rows_to_drop_indices)} filas especificadas.")
        
        # 5. Añadir nuevas columnas (Ahora más simple, sin necesidad de letras)
        if 'COLS_TO_ADD' in selected_template and selected_template['COLS_TO_ADD']:
            items_to_add = [item.strip() for item in selected_template['COLS_TO_ADD'].split(',')]
            
            for item in items_to_add:
                # Extraemos el nombre antes del paréntesis si existe, o el nombre directo
                name = item.split('(')[0].strip()
                if name not in df_new.columns:
                    df_new[name] = "" # Crea la columna al final
                    print(f"Columna creada: '{name}'")

        # 6. Reordenamiento Final Dinámico
        if 'ORDERED_COLS' in selected_template and selected_template['ORDERED_COLS']:
            # Como ya normalizamos al leer el archivo, solo limpiamos comillas si existieran
            raw_ordered_cols = selected_template['ORDERED_COLS'].replace("'", "").replace('"', "")
            ordered_cols = [col.strip() for col in raw_ordered_cols.split(',')]
            
            # Caso especial: Si el Excel usa "concepto" pero tu orden pide "descripcion"
            # (Ambos ya están en minúsculas por la normalización previa)
            if 'concepto' in df_new.columns and 'descripcion' in ordered_cols:
                df_new.rename(columns={'concepto': 'descripcion'}, inplace=True)
            
            # Ahora el cruce de datos es exacto: minúscula vs minúscula
            existing_cols = [col for col in ordered_cols if col in df_new.columns]
            
            df_new = df_new[existing_cols]
            print(f"Columnas reordenadas exitosamente.")

        # Extraemos los nombres exactos definidos en el template.txt
        # Si no existen, usamos valores por defecto seguros
        source_name = selected_template.get('SOURCE_COL', '')
        type_name = selected_template.get('TYPE_COL', '')
        cat_name = selected_template.get('CAT_COL', '')

        print(f"\n¡Éxito! Template aplicado. Columnas: {source_name}, {type_name}, {cat_name}")
                
        # IMPORTANTE: Devolvemos el DataFrame, el Template y la identidad de las columnas
        return df_new, selected_template, (source_name, type_name, cat_name)

    except Exception as e:
        print(f"Ocurrió un error al aplicar el formato: {e}")
        return df, None, (None, None, None)    
    
def main():
    df = None
    file_path = ""
    while True:
        file_path = input("Por favor, ingresa la ruta completa de tu archivo Excel: ").strip()
        if (file_path.startswith(("'", '"')) and file_path.endswith(("'", '"')) 
                and file_path[0] == file_path[-1]):
            file_path = file_path[1:-1]
        try:
            df = pd.read_excel(file_path)
            print("\n¡Archivo Excel cargado con éxito!")
            break
        except Exception as e:
            print(f"Error al leer el archivo: {e}")

    while True:
        print("\n--- Menú Principal ---")
        print("1. Usar un template (Categorizar + Limpiar)")
        print("2. Categorizar registros (Manual)")
        print("S. Guardar y salir")
        print("F. Salir sin guardar")
        
        choice = input("\nElige una opción: ").upper()
        
        if choice == '1':
            # Recibimos DF, Template y la identidad de columnas
            df, template, cols_info = apply_format(df, file_path)
            
            if df is not None:
                # Pasamos la identidad al categorizador
                df, final_src, final_cat = categorize(df, template, cols_info)
                # El aprendizaje usa los nombres resultantes
                learn_and_suggest(df, final_src, final_cat)
                
        elif choice == '2': #Esta opcion se esta quedando obsoleta
            # Al no haber template, categorize pedirá los nombres manualmente
            df, final_src, final_cat = categorize(df, None)
            learn_and_suggest(df, final_src, final_cat)

        elif choice == 'S':
            try:
                new_file_path = os.path.splitext(file_path)[0] + "_modificado.xlsx"
                df.to_excel(new_file_path, index=False)
                print(f"\nArchivo guardado exitosamente como '{new_file_path}'!")
                break
            except Exception as e:
                print(f"Error al guardar: {e}")
        elif choice == 'F':
            print("Saliendo sin guardar cambios.")
            break

if __name__ == "__main__":
    main()