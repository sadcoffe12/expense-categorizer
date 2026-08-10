"""
Template and Excel format handling for expense categorization.
Manages template definitions and applies formatting to Excel files.
"""

import os
import openpyxl
import pandas as pd
import config
from categorizer import text_normalizer


def ensure_templates_file_valid() -> bool:
    """
    Verify and recreate templates.txt if missing or empty.
    
    Creates a default template example for users to customize.
    
    Returns:
        True if file was recreated, False if it already existed
    """
    if not os.path.exists(config.TEMPLATES_FILE) or os.path.getsize(config.TEMPLATES_FILE) == 0:
        default_template = """TEMPLATE_NAME: My Template
HEADER_ROW: 1
START_ROW: 2
START_COL: A
END_COL: Z
COLS_TO_DROP: 
ROWS_TO_DROP: 
COLS_TO_ADD: 
ORDERED_COLS: 
SOURCE_COL: Descripcion
TYPE_COL: Tipo
CAT_COL: Categoria
"""
        with open(config.TEMPLATES_FILE, "w", encoding="utf-8") as f:
            f.write(default_template)
        print(f"ℹ️  File '{config.TEMPLATES_FILE}' created with example template.")
        return True
    return False


def load_templates() -> dict:
    """
    Load all templates from templates.txt file.
    
    Returns:
        Dictionary mapping template names to template configuration dicts
    """
    templates = {}
    
    try:
        with open(config.TEMPLATES_FILE, "r") as f:
            content = f.read().strip()
            if not content:
                print("Templates file is empty. Using default example.")
                return {}
            
            # Split by "---" separator
            blocks = content.split("---")
            
            for block in blocks:
                block = block.strip()
                if not block:
                    continue
                
                template_dict = {}
                for line in block.split('\n'):
                    line = line.strip()
                    if ':' in line:
                        try:
                            key, value = line.split(": ", 1)
                            key = key.strip()
                            value = value.strip()
                            
                            # Normalize certain fields
                            keys_to_normalize = ['COLS_TO_ADD', 'ORDERED_COLS', 'SOURCE_COL', 'TYPE_COL', 'CAT_COL']
                            if key in keys_to_normalize:
                                if ',' in value:
                                    value = ", ".join([text_normalizer.normalize_text(v) for v in value.split(',')])
                                else:
                                    value = text_normalizer.normalize_text(value)
                            
                            template_dict[key] = value
                        except ValueError:
                            continue
                
                if 'TEMPLATE_NAME' in template_dict:
                    templates[template_dict['TEMPLATE_NAME']] = template_dict
                    
    except Exception as e:
        print(f"Error reading templates: {e}")
        return {}

    return templates


def apply_format(df: pd.DataFrame, file_path: str) -> tuple:
    """
    Apply template formatting to DataFrame from Excel file.
    
    Reads data from specified Excel range, applies formatting rules:
    - Normalize column headers
    - Drop specified columns/rows
    - Add new columns
    - Reorder columns as specified
    
    Returns the formatted DataFrame in memory (does not save file).
    
    Args:
        df: Current DataFrame (will be replaced with formatted version)
        file_path: Path to Excel file to load and format
        
    Returns:
        Tuple of (formatted_df, template_dict, (source_col, type_col, category_col))
    """
    print("--- Apply Format Template ---")
    
    ensure_templates_file_valid()
    
    if not os.path.exists(config.TEMPLATES_FILE):
        print("No format templates found.")
        return df, None, (None, None, None)

    # Load available templates
    templates = load_templates()

    if not templates:
        print("No format templates found.")
        return df, None, (None, None, None)

    print("Available format templates:")
    for i, name in enumerate(templates.keys(), 1):
        print(f"  {i}. {name}")
    
    try:
        choice = int(input("Select template number: "))
        selected_name = list(templates.keys())[choice - 1]
        selected_template = templates[selected_name]
    except (ValueError, IndexError):
        print("Invalid selection.")
        return df, None, (None, None, None)

    # Apply the selected template
    try:
        # Load Excel workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Parse template parameters
        header_row = int(selected_template.get('HEADER_ROW', 1))
        start_row = int(selected_template.get('START_ROW', 2))
        start_col_letter = selected_template.get('START_COL', 'A')
        end_col_letter = selected_template.get('END_COL', 'Z')
        
        start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)
        end_col_idx = openpyxl.utils.column_index_from_string(end_col_letter)

        # Read data from Excel range
        data = []
        for r in range(header_row, ws.max_row + 1):
            row_data = [ws.cell(row=r, column=c).value for c in range(start_col_idx, end_col_idx + 1)]
            data.append(row_data)

        # Create DataFrame in memory
        if not data:
            print("No data found in specified range.")
            return df, None, (None, None, None)
            
        df_new = pd.DataFrame(data[1:], columns=data[0])
        
        # Normalize headers (remove accents and special chars)
        df_new.columns = [text_normalizer.normalize_text(col) for col in df_new.columns]
        print("Headers normalized (accents and symbols removed).")

        # Drop completely empty rows
        df_new.dropna(how='all', inplace=True)

        # 3. Drop specified columns
        if 'COLS_TO_DROP' in selected_template and selected_template['COLS_TO_DROP']:
            cols_to_drop = [col.strip() for col in selected_template['COLS_TO_DROP'].split(',') if col.strip()]
            headers_map = {openpyxl.utils.get_column_letter(start_col_idx + i): header 
                          for i, header in enumerate(data[0])}
            cols_to_drop_names = [headers_map[letter] for letter in cols_to_drop if letter in headers_map]
            cols_to_drop_names_existing = [name for name in cols_to_drop_names if name in df_new.columns]
            
            df_new.drop(columns=cols_to_drop_names_existing, inplace=True, errors='ignore')
            if cols_to_drop_names_existing:
                print(f"Columns dropped: {', '.join(cols_to_drop_names_existing)}")

        # 4. Drop specified rows
        if 'ROWS_TO_DROP' in selected_template and selected_template['ROWS_TO_DROP']:
            rows_to_drop_str = [r.strip() for r in selected_template['ROWS_TO_DROP'].split(',') if r.strip()]
            rows_to_drop_indices = [int(r_str) - start_row for r_str in rows_to_drop_str]
            rows_to_drop_indices = [idx for idx in rows_to_drop_indices if 0 <= idx < len(df_new)]
            
            df_new.drop(index=rows_to_drop_indices, inplace=True, errors='ignore')
            if rows_to_drop_indices:
                print(f"Dropped {len(rows_to_drop_indices)} specified rows.")
        
        # 5. Add new columns
        if 'COLS_TO_ADD' in selected_template and selected_template['COLS_TO_ADD']:
            items_to_add = [item.strip() for item in selected_template['COLS_TO_ADD'].split(',')]
            
            for item in items_to_add:
                name = item.split('(')[0].strip()
                if name not in df_new.columns:
                    df_new[name] = ""
                    print(f"Column created: '{name}'")

        # 6. Reorder columns (flexible matching)
        if 'ORDERED_COLS' in selected_template and selected_template['ORDERED_COLS']:
            raw_ordered_cols = selected_template['ORDERED_COLS'].replace("'", "").replace('"', "")
            ordered_cols = [col.strip() for col in raw_ordered_cols.split(',')]
            
            # Create flexible mapping: normalized_name -> original_column
            col_map = {}
            for df_col in df_new.columns:
                normalized_df_col = text_normalizer.normalize_col_name(df_col)
                col_map[normalized_df_col] = df_col
            
            # Find which template columns exist in DataFrame
            existing_cols = []
            for template_col in ordered_cols:
                normalized_template_col = text_normalizer.normalize_col_name(template_col)
                if normalized_template_col in col_map:
                    existing_cols.append(col_map[normalized_template_col])
            
            if existing_cols:
                df_new = df_new[existing_cols]
                print("Columns reordered successfully.")
            else:
                print("Warning: No matching columns found for reordering.")

        # Extract column names from template
        source_name = selected_template.get('SOURCE_COL', '')
        type_name = selected_template.get('TYPE_COL', '')
        cat_name = selected_template.get('CAT_COL', '')

        print(f"\n✅ Template applied. Columns: {source_name}, {type_name}, {cat_name}")
                
        return df_new, selected_template, (source_name, type_name, cat_name)

    except Exception as e:
        print(f"Error applying format: {e}")
        return df, None, (None, None, None)
